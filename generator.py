"""
TON Wallet Generator - Hana
Tạo ví TON (24 từ mnemonic) và xuất ra file Excel

"""

import os
import sys
from datetime import datetime

def generate_wallets(count):
    """Tạo ví TON và trả về danh sách (mnemonic, address)"""
    try:
        from pytoniq_core.crypto.keys import mnemonic_new, mnemonic_to_private_key
        from pytoniq_core import Address, StateInit, begin_cell, Cell
        from tonutils.wallet import WalletV5R1
    except ImportError:
        print("Thiếu thư viện! Chạy lệnh sau để cài đặt:")
        print("pip install pytoniq-core tonutils openpyxl")
        return None
    
    wallets = []
    
    for i in range(count):
        # Tạo mnemonic 24 từ
        mnemonic = mnemonic_new(24)
        mnemonic_str = ' '.join(mnemonic)
        
        # Tạo địa chỉ ví từ mnemonic (WalletV5R1)
        try:
            pub_key, priv_key = mnemonic_to_private_key(mnemonic)
            
            # WalletV5R1 code và data
            wallet_code = Cell.one_from_boc(bytes.fromhex(WalletV5R1.CODE_HEX))
            wallet_id = 2147483409  # mainnet WalletV5R1
            
            wallet_data = (
                begin_cell()
                .store_bit(1)  # is_signature_allowed = true
                .store_uint(0, 32)  # seqno = 0
                .store_int(wallet_id, 32)  # wallet_id
                .store_bytes(pub_key)  # public_key (256 bits)
                .store_bit(0)  # empty extensions dict
                .end_cell()
            )
            
            state_init = StateInit(code=wallet_code, data=wallet_data)
            address = Address((0, state_init.serialize().hash))
            # Non-bounceable (UQ...) - dạng thông thường để chuyển/nhận
            address_str = address.to_str(is_bounceable=False, is_url_safe=True)
            
            wallets.append({
                'index': i + 1,
                'mnemonic': mnemonic_str,
                'address': address_str
            })
            
            print(f"Đã tạo ví {i + 1}/{count}: {address_str}")
            
        except Exception as e:
            print(f"Lỗi tạo ví {i + 1}: {e}")
            continue
    
    return wallets

def export_to_excel(wallets, filename):
    """Xuất danh sách ví ra file Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        print("Thiếu thư viện openpyxl! Chạy: pip install openpyxl")
        return False
    
    wb = Workbook()
    ws = wb.active
    ws.title = "TON Wallets"
    
    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['STT', 'Mnemonic (24 từ)', 'Địa chỉ ví (UQ)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data
    for row, wallet in enumerate(wallets, 2):
        ws.cell(row=row, column=1, value=wallet['index']).border = border
        ws.cell(row=row, column=2, value=wallet['mnemonic']).border = border
        ws.cell(row=row, column=3, value=wallet['address']).border = border
    
    # Điều chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 120
    ws.column_dimensions['C'].width = 55
    
    # Lưu file
    wb.save(filename)
    return True

def main():
    print("="*60)
    print("       TON WALLET GENERATOR - HANA")
    print("       Tạo ví TON (24 từ) ")
    print("="*60)
    print()
    
    # Hỏi số lượng
    while True:
        try:
            count = int(input("Nhập số lượng ví cần tạo: "))
            if count <= 0:
                print("Số lượng phải > 0!")
                continue
            break
        except ValueError:
            print("Vui lòng nhập số!")
    
    print()
    print(f"Đang tạo {count} ví TON...")
    print("-"*60)
    
    # Tạo ví
    wallets = generate_wallets(count)
    
    if not wallets:
        print("Không tạo được ví nào!")
        return
    
    print("-"*60)
    print(f"Đã tạo thành công {len(wallets)} ví!")
    print()
    
    # Xuất Excel
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"ton_wallets_{timestamp}.xlsx"
    
    print(f"Đang xuất ra file Excel: {filename}")
    
    if export_to_excel(wallets, filename):
        print(f"✓ Đã xuất thành công: {filename}")
        
        # Xuất thêm file txt chỉ chứa mnemonic 
        txt_filename = f"ton_wallets_{timestamp}.txt"
        with open(txt_filename, 'w', encoding='utf-8') as f:
            for wallet in wallets:
                f.write(wallet['mnemonic'] + '\n')
        print(f"✓ Đã xuất file mnemonic: {txt_filename}")
    else:
        print("✗ Lỗi xuất file Excel!")
    
    print()
    print("="*60)

if __name__ == '__main__':
    main()
